import io
import json
from datetime import date
from collections import defaultdict
from dateutil.relativedelta import relativedelta

from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Q, F
from django.core.paginator import Paginator
from dotenv import load_dotenv

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404


from .models import RegistroRefeicao, TabelaPreco, LocalRefeicao, Fazenda, Perfil
from .forms import RegistroRefeicaoForm, TabelaPrecoForm

from django.db.models.functions import TruncMonth
from django.db.models import Sum, Q, F, Case, When, DecimalField, IntegerField

load_dotenv()


@login_required
def painel_refeicoes(request):
    formato_clicado = request.GET.get('formato', 'filtrar')

    if formato_clicado == 'pdf':
        return exportar_pdf(request)
    elif formato_clicado == 'excel':
        return exportar_refeicoes_excel(request)

    registros = RegistroRefeicao.objects.select_related('fazenda').all().order_by('-data_consumo', '-id')

    is_dono = True
    fazenda_usuario = None
    if hasattr(request.user, 'perfil'):
        is_dono = request.user.perfil.is_dono
        if request.user.perfil.fazenda_lotacao:
            fazenda_usuario = request.user.perfil.fazenda_lotacao

    if not is_dono and fazenda_usuario:
        registros = registros.filter(fazenda=fazenda_usuario)

    nome_fazenda_atual = fazenda_usuario.nome if fazenda_usuario else "Todas as Fazendas"

    # Define dinamicamente as cantinas
    nome_f = fazenda_usuario.nome if fazenda_usuario else ""
    if is_dono:
        cantinas_disponiveis = list(LocalRefeicao.choices)
    else:
        if nome_f == 'Fazenda BC':
            cantinas_disponiveis = [(LocalRefeicao.CANTINA_BC, 'Cantina BC')]
        elif nome_f == 'Fazenda Matão':
            cantinas_disponiveis = [(LocalRefeicao.CANTINA_MATAO, 'Cantina Matão')]
        elif nome_f == 'Fazenda Lagoa':
            cantinas_disponiveis = [(LocalRefeicao.CANTINA_LAGOA, 'Cantina Lagoa')]
        else:
            cantinas_disponiveis = [
                ('', 'Todas'),
                (LocalRefeicao.SEDE, 'Cantina Sede'),
                (LocalRefeicao.SECADOR, 'Cantina Secador')
            ]

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    local_busca = request.GET.get('local')
    setor_busca = request.GET.get('setor')
    fazenda_busca = request.GET.get('fazenda')

    if is_dono and fazenda_busca:
        registros = registros.filter(fazenda_id=fazenda_busca)

    if not data_inicio and not data_fim:
        hoje = date.today()
        registros = registros.filter(
            data_consumo__year=hoje.year,
            data_consumo__month=hoje.month
        )
    else:
        if data_inicio:
            registros = registros.filter(data_consumo__gte=data_inicio)
        if data_fim:
            registros = registros.filter(data_consumo__lte=data_fim)

    if not is_dono and nome_f in ['Fazenda BC', 'Fazenda Matão', 'Fazenda Lagoa']:
        if nome_f == 'Fazenda BC':
            local_busca = LocalRefeicao.CANTINA_BC
        elif nome_f == 'Fazenda Matão':
            local_busca = LocalRefeicao.CANTINA_MATAO
        elif nome_f == 'Fazenda Lagoa':
            local_busca = LocalRefeicao.CANTINA_LAGOA

    if local_busca:
        registros = registros.filter(local=local_busca)
    if setor_busca:
        registros = registros.filter(setor__icontains=setor_busca)


    agregados = registros.aggregate(
        q_cafe=Sum('qtd_cafe'),
        q_buffet=Sum('qtd_almoco_buffet'),
        q_marmita=Sum('qtd_almoco_marmita'),
        q_janta=Sum('qtd_janta'),
        q_lanche=Sum('qtd_lanche'),

        v_cafe=Sum(F('qtd_cafe') * F('valor_cafe')),
        v_buffet=Sum(F('qtd_almoco_buffet') * F('valor_almoco')),
        v_marmita=Sum(F('qtd_almoco_marmita') * F('valor_almoco_marmita')),
        v_janta=Sum(F('qtd_janta') * F('valor_janta')),
        v_lanche=Sum(F('qtd_lanche') * F('valor_lanche')),

        # Puxamos a coluna que o banco já calculou
        soma_geral=Sum('valor_total')
    )

    total_gasto = agregados['soma_geral'] or 0

    total_refeicoes = (
            (agregados['q_cafe'] or 0) + (agregados['q_buffet'] or 0) +
            (agregados['q_marmita'] or 0) + (agregados['q_janta'] or 0) + (agregados['q_lanche'] or 0)
    )

    paginator = Paginator(registros, 7)
    numero_pagina = request.GET.get('page')
    page_obj = paginator.get_page(numero_pagina)

    contexto = {
        'page_obj': page_obj,
        'total_gasto': float(total_gasto),
        'total_refeicoes': total_refeicoes,
        'total_buffet': agregados['q_buffet'] or 0,
        'total_janta': agregados['q_janta'] or 0,

        'det_q_cafe': agregados['q_cafe'] or 0,
        'det_v_cafe': float(agregados['v_cafe'] or 0),
        'det_q_buffet': agregados['q_buffet'] or 0,
        'det_v_buffet': float(agregados['v_buffet'] or 0),
        'det_q_marmita': agregados['q_marmita'] or 0,
        'det_v_marmita': float(agregados['v_marmita'] or 0),
        'det_q_janta': agregados['q_janta'] or 0,
        'det_v_janta': float(agregados['v_janta'] or 0),
        'det_q_lanche': agregados['q_lanche'] or 0,
        'det_v_lanche': float(agregados['v_lanche'] or 0),

        'nome_fazenda_atual': nome_fazenda_atual,
        'cantinas_disponiveis': cantinas_disponiveis,
        'fazendas': Fazenda.objects.all() if is_dono else [],
        'is_dono': is_dono,
        'filtros': request.GET
    }

    return render(request, 'refeicoes/painel.html', contexto)

@login_required
def dashboard_refeicoes(request):
    if hasattr(request.user, 'perfil') and not request.user.perfil.is_dono:
        return redirect('painel')

    hoje = date.today()
    if hoje.month >= 9:
        ano_inicio_safra = hoje.year
    else:
        ano_inicio_safra = hoje.year - 1

    ano_fim_safra = ano_inicio_safra + 1
    nome_safra = f"Safra {str(ano_inicio_safra)[2:]}/{str(ano_fim_safra)[2:]}"

    registros = RegistroRefeicao.objects.all()

    fazenda_id = request.GET.get('fazenda')
    nome_fazenda_atual = "Todas as Fazendas"

    if fazenda_id:
        registros = registros.filter(fazenda_id=fazenda_id)
        fazenda_selecionada = Fazenda.objects.filter(id=fazenda_id).first()
        if fazenda_selecionada:
            nome_fazenda_atual = fazenda_selecionada.nome

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if not data_inicio and not data_fim:
        data_inicio_obj = date(ano_inicio_safra, 9, 1)
        data_fim_obj = date(ano_fim_safra, 8, 31)
        registros = registros.filter(data_consumo__gte=data_inicio_obj, data_consumo__lte=data_fim_obj)
    else:
        if data_inicio:
            from datetime import datetime
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            registros = registros.filter(data_consumo__gte=data_inicio_obj)
        else:
            data_inicio_obj = date(ano_inicio_safra, 9, 1)

        if data_fim:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            registros = registros.filter(data_consumo__lte=data_fim_obj)
        else:
            data_fim_obj = date(ano_fim_safra, 8, 31)

    soma_total = registros.aggregate(
        total=Sum('valor_total'), cafe=Sum('qtd_cafe'), buffet=Sum('qtd_almoco_buffet'),
        marmita=Sum('qtd_almoco_marmita'), janta=Sum('qtd_janta'), lanche=Sum('qtd_lanche')
    )

    detalhes = registros.aggregate(
        v_cafe=Sum(F('qtd_cafe') * F('valor_cafe')),
        v_buffet=Sum(F('qtd_almoco_buffet') * F('valor_almoco')),
        v_marmita=Sum(F('qtd_almoco_marmita') * F('valor_almoco_marmita')),
        v_janta=Sum(F('qtd_janta') * F('valor_janta')),
        v_lanche=Sum(F('qtd_lanche') * F('valor_lanche'))
    )

    total_gasto = ((detalhes['v_cafe'] or 0) + (detalhes['v_buffet'] or 0) + (detalhes['v_marmita'] or 0) + (
                detalhes['v_janta'] or 0) + (detalhes['v_lanche'] or 0))
    total_refeicoes = ((soma_total['cafe'] or 0) + (soma_total['buffet'] or 0) + (soma_total['marmita'] or 0) + (
                soma_total['janta'] or 0) + (soma_total['lanche'] or 0))


    expressao_qtd = F('qtd_cafe') + F('qtd_almoco_buffet') + F('qtd_almoco_marmita') + F('qtd_janta') + F('qtd_lanche')
    filtro_terceiros = Q(setor__icontains='terceirizado') | Q(setor__icontains='terceiros')

    dados_agrupados_sql = registros.annotate(
        mes_exato=TruncMonth('data_consumo')
    ).values('mes_exato').annotate(
        terc_v=Sum(Case(When(filtro_terceiros, then=F('valor_total')), default=0, output_field=DecimalField())),
        colab_v=Sum(Case(When(~filtro_terceiros, then=F('valor_total')), default=0, output_field=DecimalField())),
        terc_q=Sum(Case(When(filtro_terceiros, then=expressao_qtd), default=0, output_field=IntegerField())),
        colab_q=Sum(Case(When(~filtro_terceiros, then=expressao_qtd), default=0, output_field=IntegerField()))
    ).order_by('mes_exato')

    # Povoa o mapa para o gráfico ler
    mapa_meses = {}
    for linha in dados_agrupados_sql:
        if linha['mes_exato']:
            ano = linha['mes_exato'].year
            mes = linha['mes_exato'].month
            mapa_meses[(ano, mes)] = {
                'colab_v': float(linha['colab_v'] or 0),
                'colab_q': linha['colab_q'] or 0,
                'terc_v': float(linha['terc_v'] or 0),
                'terc_q': linha['terc_q'] or 0
            }

    total_colab_periodo = sum(mes['colab_v'] for mes in mapa_meses.values())
    total_terc_periodo = sum(mes['terc_v'] for mes in mapa_meses.values())

    # 6. Preparar Listas para o Gráfico (Instantâneo)
    meses_labels = []
    dados_colaboradores = []
    dados_terceirizados = []
    qtds_colaboradores = []
    qtds_terceirizados = []

    nomes_meses = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

    data_atual_iter = data_inicio_obj.replace(day=1)
    data_limite_iter = data_fim_obj.replace(day=1)

    while data_atual_iter <= data_limite_iter:
        ano_alvo = data_atual_iter.year
        mes_alvo = data_atual_iter.month

        label = f"{nomes_meses[mes_alvo]}/{str(ano_alvo)[2:]}"
        meses_labels.append(label)

        dados_do_mes = mapa_meses.get((ano_alvo, mes_alvo), {'colab_v': 0, 'colab_q': 0, 'terc_v': 0, 'terc_q': 0})

        dados_colaboradores.append(float(dados_do_mes['colab_v']))
        dados_terceirizados.append(float(dados_do_mes['terc_v']))
        qtds_colaboradores.append(dados_do_mes['colab_q'])
        qtds_terceirizados.append(dados_do_mes['terc_q'])

        data_atual_iter += relativedelta(months=1)
        if len(meses_labels) >= 24: break

    contexto = {
        'total_gasto': float(total_gasto), 'total_refeicoes': total_refeicoes,
        'total_buffet': soma_total['buffet'] or 0, 'total_janta': soma_total['janta'] or 0,
        'det_q_cafe': soma_total['cafe'] or 0, 'det_v_cafe': float(detalhes['v_cafe'] or 0),
        'det_q_buffet': soma_total['buffet'] or 0, 'det_v_buffet': float(detalhes['v_buffet'] or 0),
        'det_q_marmita': soma_total['marmita'] or 0, 'det_v_marmita': float(detalhes['v_marmita'] or 0),
        'det_q_janta': soma_total['janta'] or 0, 'det_v_janta': float(detalhes['v_janta'] or 0),
        'det_q_lanche': soma_total['lanche'] or 0, 'det_v_lanche': float(detalhes['v_lanche'] or 0),

        'total_colab_periodo': float(total_colab_periodo),
        'total_terc_periodo': float(total_terc_periodo),
        'meses_labels': json.dumps(meses_labels),
        'dados_colaboradores': json.dumps(dados_colaboradores),
        'dados_terceirizados': json.dumps(dados_terceirizados),
        'qtds_colaboradores': json.dumps(qtds_colaboradores),
        'qtds_terceirizados': json.dumps(qtds_terceirizados),

        'nome_safra': nome_safra,
        'nome_fazenda_atual': nome_fazenda_atual,
        'fazendas_disponiveis': Fazenda.objects.all(),
        'filtros': request.GET
    }
    return render(request, 'refeicoes/dashboard.html', contexto)

@login_required
def novo_registro(request):
    if hasattr(request.user, 'perfil') and request.user.perfil.is_dono:
        return redirect('painel')

    if request.method == "POST":
        form = RegistroRefeicaoForm(request.POST, usuario=request.user)
        if form.is_valid():
            registro = form.save(commit=False)

            if hasattr(request.user, 'perfil') and request.user.perfil.fazenda_lotacao:
                registro.fazenda = request.user.perfil.fazenda_lotacao

            registro.save()
            return redirect('painel')
    else:
        form = RegistroRefeicaoForm(usuario=request.user)

    return render(request, 'refeicoes/novo_registro.html', {'form': form})

@login_required
def editar_registro(request, id):
    registro = get_object_or_404(RegistroRefeicao, id=id)

    if hasattr(request.user, 'perfil'):
        if request.user.perfil.is_dono:
            return redirect('painel')

        if not request.user.perfil.is_dono and registro.fazenda != request.user.perfil.fazenda_lotacao:
            return redirect('painel')

    if request.method == 'POST':
        form = RegistroRefeicaoForm(request.POST, instance=registro, usuario=request.user)
        if form.is_valid():
            form.save()
            return redirect('painel')
    else:
        form = RegistroRefeicaoForm(instance=registro, usuario=request.user)
    return render(request, 'refeicoes/novo_registro.html', {'form': form, 'registro': registro})

@login_required
def excluir_registro(request, id):
    registro = get_object_or_404(RegistroRefeicao, id=id)

    if hasattr(request.user, 'perfil'):
        if request.user.perfil.is_dono:
            return redirect('painel')

        if not request.user.perfil.is_dono and registro.fazenda != request.user.perfil.fazenda_lotacao:
            return redirect('painel')

    registro.delete()
    return redirect('painel')

@login_required
def exportar_pdf(request):
    from reportlab.platypus import PageBreak

    registros = RegistroRefeicao.objects.all()

    is_dono = True
    fazenda_usuario = None
    if hasattr(request.user, 'perfil'):
        is_dono = request.user.perfil.is_dono
        if request.user.perfil.fazenda_lotacao:
            fazenda_usuario = request.user.perfil.fazenda_lotacao

    if not is_dono and fazenda_usuario:
        registros = registros.filter(fazenda=fazenda_usuario)

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    fazenda_busca = request.GET.get('fazenda')

    if is_dono and fazenda_busca:
        registros = registros.filter(fazenda_id=fazenda_busca)

    if not data_inicio and not data_fim:
        hoje = date.today()
        registros = registros.filter(data_consumo__year=hoje.year, data_consumo__month=hoje.month)
    else:
        if data_inicio: registros = registros.filter(data_consumo__gte=data_inicio)
        if data_fim: registros = registros.filter(data_consumo__lte=data_fim)

    local_busca = request.GET.get('local')
    setor_busca = request.GET.get('setor')

    if local_busca: registros = registros.filter(local=local_busca)
    if setor_busca: registros = registros.filter(setor__icontains=setor_busca)

    registros = registros.order_by('fazenda__nome', 'setor', 'data_consumo')

    dados_agrupados = defaultdict(lambda: defaultdict(list))
    total_geral = 0

    for r in registros:
        nome_setor = r.get_setor_display() if hasattr(r, 'get_setor_display') else r.setor

        if is_dono:
            nome_fazenda = r.fazenda.nome if r.fazenda else "Sem Fazenda"
            dados_agrupados[nome_fazenda][nome_setor].append(r)
        else:
            dados_agrupados[""][nome_setor].append(r)

        total_geral += float(r.valor_total or 0)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40,
                            bottomMargin=40)
    elementos = []
    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle('TituloModerno', parent=estilos['Heading1'], alignment=TA_CENTER,
                                   fontSize=20, textColor=colors.black, spaceAfter=5, fontName='Helvetica-Bold')
    estilo_subtitulo = ParagraphStyle('Subtitulo', parent=estilos['Normal'], alignment=TA_CENTER,
                                      fontSize=11, textColor=colors.black, spaceAfter=25)
    estilo_nome_fazenda = ParagraphStyle('NomeFazenda', parent=estilos['Heading1'], fontSize=16,
                                         textColor=colors.HexColor('#1f2937'), spaceBefore=25, spaceAfter=10,
                                         fontName='Helvetica-Bold', alignment=TA_CENTER)
    estilo_nome_setor = ParagraphStyle('NomeSetor', parent=estilos['Heading2'], fontSize=14,
                                textColor=colors.black,spaceBefore=15, spaceAfter=10, fontName='Helvetica-Bold')

    elementos.append(Paragraph("Relatório de Refeições", estilo_titulo))
    elementos.append(Paragraph("Extrato analítico gerado pelo sistema.", estilo_subtitulo))

    def formata_rs(valor):
        if valor and valor > 0:
            return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return '-'

    is_primeira_fazenda = True

    for fazenda_nome, setores in dados_agrupados.items():
        if not is_primeira_fazenda:
            elementos.append(PageBreak())

        if fazenda_nome != "":
            estilo_usado = estilo_nome_fazenda if is_primeira_fazenda else ParagraphStyle('NomeFazendaNova',
                                                                    parent=estilo_nome_fazenda, spaceBefore=0)
            elementos.append(Paragraph(f"FAZENDA: {fazenda_nome.upper()}", estilo_usado))

        is_primeira_fazenda = False

        for setor_nome, lista_refeicoes in setores.items():
            elementos.append(Paragraph(f"Setor: {setor_nome}", estilo_nome_setor))
            cabecalho = ['Data', 'Cantina', 'Café', 'Buffet', 'Marm.', 'Janta', 'Lanche', 'Valor Total']
            dados_tabela = [cabecalho]

            total_deste_setor = 0
            v_cafe = v_buffet = v_marmita = v_janta = v_lanche = 0

            for r in lista_refeicoes:
                # Mantido apenas para o subtotal do rodapé do PDF
                c_v = float(r.qtd_cafe * r.valor_cafe)
                b_v = float(r.qtd_almoco_buffet * r.valor_almoco)
                m_v = float(r.qtd_almoco_marmita * r.valor_almoco_marmita)
                j_v = float(r.qtd_janta * r.valor_janta)
                l_v = float(r.qtd_lanche * r.valor_lanche)

                v_cafe += c_v
                v_buffet += b_v
                v_marmita += m_v
                v_janta += j_v
                v_lanche += l_v

                valor_linha = float(r.valor_total or 0)
                total_deste_setor += valor_linha

                linha = [
                    r.data_formatada() if hasattr(r, 'data_formatada') else r.data_consumo.strftime('%d/%m/%Y'),
                    r.get_local_display() if hasattr(r, 'get_local_display') else r.local,
                    r.qtd_cafe or '-', r.qtd_almoco_buffet or '-', r.qtd_almoco_marmita or '-',
                    r.qtd_janta or '-', r.qtd_lanche or '-',
                    formata_rs(valor_linha)
                ]
                dados_tabela.append(linha)

            linha_total = [
                '', 'SUBTOTAL DO SETOR:', formata_rs(v_cafe), formata_rs(v_buffet),
                formata_rs(v_marmita), formata_rs(v_janta), formata_rs(v_lanche), formata_rs(total_deste_setor)
            ]
            dados_tabela.append(linha_total)

            estilo_tabela_minimalista = TableStyle([
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-2, -1), 'CENTER'),
                ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
                ('LINEBELOW', (0, 0), (-1, 0), 1.2, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (2, -1), (-2, -1), 9),
            ])

            tabela = Table(dados_tabela, colWidths=[70, 160, 65, 65, 65, 65, 65, 90], repeatRows=1)
            tabela.setStyle(estilo_tabela_minimalista)
            elementos.append(tabela)
            elementos.append(Spacer(1, 20))

    elementos.append(Spacer(1, 10))
    estilo_total_geral = ParagraphStyle('TotalGeral', parent=estilos['Heading2'], alignment=TA_RIGHT,
                                    textColor=colors.black, spaceTop=10, fontName='Helvetica-Bold', fontSize=14)
    texto_total_geral = (f"CUSTO TOTAL DO PERÍODO: R$ {total_geral:,.2f}"
                         .replace(',', 'X').replace('.', ',').replace('X', '.'))
    elementos.append(Paragraph(texto_total_geral, estilo_total_geral))

    doc.build(elementos)
    buffer.seek(0)

    hoje = date.today()
    nome_arquivo = f"Relatorio_Refeicoes_{hoje.strftime('%m_%Y')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=nome_arquivo)

@login_required
def exportar_refeicoes_excel(request):
    import openpyxl
    from openpyxl.styles import Font
    from django.http import HttpResponse
    from collections import defaultdict
    from datetime import date

    registros = RegistroRefeicao.objects.all()

    is_dono = True
    fazenda_usuario = None
    if hasattr(request.user, 'perfil'):
        is_dono = request.user.perfil.is_dono
        if request.user.perfil.fazenda_lotacao:
            fazenda_usuario = request.user.perfil.fazenda_lotacao

    if not is_dono and fazenda_usuario:
        registros = registros.filter(fazenda=fazenda_usuario)

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    fazenda_busca = request.GET.get('fazenda')

    if is_dono and fazenda_busca:
        registros = registros.filter(fazenda_id=fazenda_busca)

    if not data_inicio and not data_fim:
        hoje = date.today()
        registros = registros.filter(data_consumo__year=hoje.year, data_consumo__month=hoje.month)
    else:
        if data_inicio: registros = registros.filter(data_consumo__gte=data_inicio)
        if data_fim: registros = registros.filter(data_consumo__lte=data_fim)

    local_busca = request.GET.get('local')
    setor_busca = request.GET.get('setor')

    if local_busca: registros = registros.filter(local=local_busca)
    if setor_busca: registros = registros.filter(setor__icontains=setor_busca)

    registros = registros.order_by('fazenda__nome', 'data_consumo', 'setor')

    dados_agrupados = defaultdict(list)
    for r in registros:
        if is_dono:
            nome_fazenda = r.fazenda.nome if r.fazenda else "Sem Fazenda"
        else:
            nome_fazenda = fazenda_usuario.nome if fazenda_usuario else "Lançamentos"

        nome_aba = str(nome_fazenda)[:31]
        dados_agrupados[nome_aba].append(r)

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    if not dados_agrupados:
        wb.create_sheet(title="Sem Dados")

    for nome_aba, lista_refeicoes in dados_agrupados.items():
        ws = wb.create_sheet(title=nome_aba)

        cabecalho = ['Data', 'Local', 'Setor', 'Café', 'Almoço Buffet',
                     'Almoço Marmita', 'Janta', 'Lanche', 'Valor Total']
        ws.append(cabecalho)

        for col_num in range(1, len(cabecalho) + 1):
            ws.cell(row=1, column=col_num).font = Font(bold=True)

        for r in lista_refeicoes:
            valor_linha = float(r.valor_total or 0)

            linha = [
                r.data_formatada() if hasattr(r, 'data_formatada') else r.data_consumo.strftime('%d/%m/%Y'),
                r.get_local_display() if hasattr(r, 'get_local_display') else r.local,
                r.get_setor_display() if hasattr(r, 'get_setor_display') else r.setor,
                r.qtd_cafe or 0,
                r.qtd_almoco_buffet or 0,
                r.qtd_almoco_marmita or 0,
                r.qtd_janta or 0,
                r.qtd_lanche or 0,
                valor_linha
            ]
            ws.append(linha)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['I'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    hoje_str = date.today().strftime('%m_%Y')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Refeicoes_{hoje_str}.xlsx"'

    wb.save(response)
    return response

@login_required
def configurar_precos(request):
    fazenda_usuario = None
    if hasattr(request.user, 'perfil') and request.user.perfil.fazenda_lotacao:
        fazenda_usuario = request.user.perfil.fazenda_lotacao

    if fazenda_usuario:
        tabela = TabelaPreco.objects.filter(fazenda=fazenda_usuario).first()
        if not tabela:
            tabela = TabelaPreco.objects.create(fazenda=fazenda_usuario)
    else:
        tabela = TabelaPreco.objects.filter(fazenda__isnull=True).first()
        if not tabela:
            tabela = TabelaPreco.objects.create()

    if request.method == 'POST':
        form = TabelaPrecoForm(request.POST, instance=tabela)
        if form.is_valid():
            form.save()
            return redirect('painel')
    else:
        form = TabelaPrecoForm(instance=tabela)

    contexto = {
        'form': form,
        'tabela': tabela,
        'nome_fazenda': fazenda_usuario.nome if fazenda_usuario else "Matriz (Padrão)"
    }
    return render(request, 'refeicoes/configurar_precos.html', contexto)

@login_required
def gerenciar_usuarios(request):
    is_dono = request.user.is_superuser or (hasattr(request.user, 'perfil') and request.user.perfil.is_dono)

    if not is_dono:
        return redirect('painel')

    usuarios = list(User.objects.all().select_related('perfil'))

    for u in usuarios:
        u.is_logged_in_user = (u.id == request.user.id)
        u.is_admin_flag = u.is_superuser or (hasattr(u, 'perfil') and u.perfil.is_dono)


    usuarios.sort(key=lambda u: (not u.is_admin_flag, u.username.lower()))

    return render(request, 'refeicoes/gerenciar_usuarios.html',
                  {'usuarios': usuarios, 'is_dono': True})

@login_required
def criar_usuario(request):
    is_dono = request.user.is_superuser or (hasattr(request.user, 'perfil') and request.user.perfil.is_dono)

    if not is_dono:
        return redirect('painel')

    fazendas = Fazenda.objects.all()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        fazenda_id = request.POST.get('fazenda')
        nivel_acesso = request.POST.get('nivel_acesso')  # Captura a escolha do Admin


        if nivel_acesso != 'admin' and not fazenda_id:
            return render(request, 'refeicoes/criar_usuario.html', {
                'fazendas': fazendas,
                'erro': 'Obrigatório: Usuários comuns devem estar vinculados a uma Fazenda.',
                'is_dono': True
            })

        if User.objects.filter(username=username).exists():
            return render(request, 'refeicoes/criar_usuario.html', {
                'fazendas': fazendas,
                'erro': 'Este nome de usuário já está em uso. Escolha outro.',
                'is_dono': True
            })

        # Cria o usuário
        novo_user = User.objects.create_user(username=username, password=password)


        if nivel_acesso == 'admin':
            fazenda = None
        else:
            fazenda = get_object_or_404(Fazenda, id=fazenda_id) if fazenda_id else None

        perfil, created = Perfil.objects.get_or_create(usuario=novo_user)
        perfil.fazenda_lotacao = fazenda
        perfil.is_dono = (nivel_acesso == 'admin')  # Se escolheu admin, fica True. Se não, False.
        perfil.save()

        return redirect('gerenciar_usuarios')

    return render(request, 'refeicoes/criar_usuario.html', {'fazendas': fazendas, 'is_dono': True})

@login_required
def excluir_usuario(request, user_id):
    is_dono = request.user.is_superuser or (hasattr(request.user, 'perfil') and request.user.perfil.is_dono)

    if not is_dono:
        return redirect('painel')

    usuario = get_object_or_404(User, id=user_id)

    if usuario.id != request.user.id:
        usuario.delete()

    return redirect('gerenciar_usuarios')
